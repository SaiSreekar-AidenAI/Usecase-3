import logging
from pathlib import Path

import chromadb
from openpyxl import load_workbook

from .config import get_settings

logger = logging.getLogger(__name__)

_collection: chromadb.Collection | None = None


def init_chroma() -> None:
    global _collection
    settings = get_settings()
    persist_dir = settings.chroma_persist_dir
    Path(persist_dir).mkdir(parents=True, exist_ok=True)

    client = chromadb.PersistentClient(path=persist_dir)
    _collection = client.get_or_create_collection(
        name="canned_responses",
        metadata={"hnsw:space": "cosine"},
    )

    if _collection.count() == 0:
        _ingest_excel(settings.excel_data_path)


def _ingest_excel(excel_path: str) -> None:
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active

    documents: list[str] = []
    metadatas: list[dict] = []
    ids: list[str] = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        if not row or not row[0]:
            continue
        category = str(row[0]).strip()
        description = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        response = str(row[2]).strip() if len(row) > 2 and row[2] else ""

        doc = f"Category: {category}\nDescription: {description}\nResponse: {response}"
        documents.append(doc)
        metadatas.append({
            "category": category,
            "description": description,
            "response": response,
        })
        ids.append(f"canned-{i}")

    if documents:
        _collection.add(documents=documents, metadatas=metadatas, ids=ids)
        logger.info(f"Ingested {len(documents)} canned responses into ChromaDB")

    wb.close()


def search_canned_responses(query: str, n_results: int = 3) -> list[dict]:
    if _collection is None:
        return []

    results = _collection.query(query_texts=[query], n_results=n_results)

    sources = []
    for i in range(len(results["ids"][0])):
        meta = results["metadatas"][0][i]
        distance = results["distances"][0][i] if results.get("distances") else 0.0
        relevance = round(1.0 - distance, 4)
        sources.append({
            "category": meta["category"],
            "description": meta["description"],
            "response": meta["response"],
            "relevance_score": relevance,
        })

    return sources
