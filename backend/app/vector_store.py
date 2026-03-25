import logging
import uuid

from google.cloud import bigquery
from google.cloud import aiplatform
from vertexai.language_models import TextEmbeddingModel
from openpyxl import load_workbook

from .config import get_settings

logger = logging.getLogger(__name__)

_bq_client: bigquery.Client | None = None
_embed_model: TextEmbeddingModel | None = None
_settings = None


def init_vector_store() -> None:
    global _bq_client, _embed_model, _settings
    _settings = get_settings()

    _bq_client = bigquery.Client(project=_settings.gcp_project_id)

    aiplatform.init(
        project=_settings.gcp_project_id,
        location=_settings.embedding_location,
    )
    _embed_model = TextEmbeddingModel.from_pretrained(_settings.embedding_model)

    # Check if table is empty and ingest if needed
    table_ref = f"{_settings.gcp_project_id}.{_settings.bq_dataset}.{_settings.bq_table}"
    query = f"SELECT COUNT(*) as cnt FROM `{table_ref}`"
    result = _bq_client.query(query).result()
    count = list(result)[0]["cnt"]

    if count == 0:
        logger.info("BigQuery table is empty, ingesting Excel data...")
        _ingest_excel(_settings.excel_data_path)
    else:
        logger.info(f"BigQuery table has {count} rows, skipping ingestion.")


def _get_embeddings(texts: list[str]) -> list[list[float]]:
    # Vertex AI batch embed (max 250 per call, chunk if needed)
    all_embeddings = []
    batch_size = 250
    for i in range(0, len(texts), batch_size):
        batch = texts[i:i + batch_size]
        embeddings = _embed_model.get_embeddings(batch)
        all_embeddings.extend([e.values for e in embeddings])
    return all_embeddings


def _ingest_excel(excel_path: str) -> None:
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active

    documents: list[str] = []
    rows_data: list[dict] = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        if not row or not row[0]:
            continue
        category = str(row[0]).strip()
        description = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        response = str(row[2]).strip() if len(row) > 2 and row[2] else ""

        doc = f"Category: {category}\nDescription: {description}\nResponse: {response}"
        documents.append(doc)
        rows_data.append({
            "id": f"canned-{i}",
            "category": category,
            "description": description,
            "response": response,
            "document": doc,
        })

    wb.close()

    if not documents:
        return

    # Get embeddings for all documents
    embeddings = _get_embeddings(documents)

    # Build rows for BigQuery insert
    table_ref = f"{_settings.gcp_project_id}.{_settings.bq_dataset}.{_settings.bq_table}"
    bq_rows = []
    for data, emb in zip(rows_data, embeddings):
        bq_rows.append({
            "id": data["id"],
            "category": data["category"],
            "description": data["description"],
            "response": data["response"],
            "document": data["document"],
            "embedding": emb,
        })

    errors = _bq_client.insert_rows_json(table_ref, bq_rows)
    if errors:
        logger.error(f"BigQuery insert errors: {errors}")
    else:
        logger.info(f"Ingested {len(bq_rows)} canned responses into BigQuery")


def search_canned_responses(query: str, n_results: int = 3) -> list[dict]:
    if _bq_client is None or _embed_model is None:
        return []

    # Embed the query
    query_embedding = _get_embeddings([query])[0]

    table_ref = f"`{_settings.gcp_project_id}.{_settings.bq_dataset}.{_settings.bq_table}`"

    # Use VECTOR_SEARCH for similarity search
    sql = f"""
        SELECT
            base.category,
            base.description,
            base.response,
            distance
        FROM VECTOR_SEARCH(
            TABLE {table_ref},
            'embedding',
            (SELECT @query_embedding AS embedding),
            top_k => @top_k,
            distance_type => 'COSINE'
        )
    """

    job_config = bigquery.QueryJobConfig(
        query_parameters=[
            bigquery.ArrayQueryParameter("query_embedding", "FLOAT64", query_embedding),
            bigquery.ScalarQueryParameter("top_k", "INT64", n_results),
        ]
    )

    results = _bq_client.query(sql, job_config=job_config).result()

    sources = []
    for row in results:
        relevance = round(1.0 - row["distance"], 4)
        sources.append({
            "category": row["category"],
            "description": row["description"],
            "response": row["response"],
            "relevance_score": relevance,
        })

    return sources
